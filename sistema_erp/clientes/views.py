
from django.shortcuts import get_object_or_404, redirect, render

from django.urls import reverse_lazy
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DetailView, View
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.mixins import LoginRequiredMixin

from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from django import forms

from .models import Cliente

# ======================================================
# FORMULARIO DE CLIENTE
# ======================================================
try:
    from .forms import ClienteForm
except Exception:
    class ClienteForm(forms.ModelForm):
        class Meta:
            model = Cliente
            fields = [
                "nombre",
                "rut",
                "direccion",
                "telefono",
                "email",
                "estadoCondicion",
            ]



# ======================================================
# EXPORTAR CLIENTES A EXCEL
# ======================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


def exportar_clientes_excel(request):
    """Exporta los clientes filtrados/ordenados a un archivo XLSX con formato profesional."""
    if Workbook is None:
        return HttpResponse(
            "openpyxl no está instalado. Agrega 'openpyxl' a requirements.txt e instala las dependencias.",
            status=500
        )

    qs = Cliente.objects.all()

    # --- Búsqueda ---
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q)
            | Q(rut__icontains=q)
            | Q(email__icontains=q)
            | Q(telefono__icontains=q)
            | Q(direccion__icontains=q)
        )

    # --- Orden ---
    sort = (request.GET.get('sort') or 'nombre').strip()
    direction = (request.GET.get('dir') or 'asc').strip().lower()
    sort_map = {
        'nombre': 'nombre',
        'rut': 'rut',
        'email': 'email',
        'telefono': 'telefono',
        'estadoCondicion': 'estadoCondicion',
        'fecha_registro': 'fecha_registro',
    }
    order_field = sort_map.get(sort, 'nombre')
    if direction == 'desc':
        order_field = f'-{order_field}'
    qs = qs.order_by(order_field)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'

    headers = [
        'ID', 'Nombre', 'RUT', 'Teléfono', 'Email', 'Dirección',
        'Estado', 'Fecha Registro'
    ]
    header_fill = PatternFill(start_color='EAF2FF', end_color='EAF2FF', fill_type='solid')
    bold = Font(bold=True, color='1f2937')
    center = Alignment(horizontal='center', vertical='center')

    ws.append(headers)
    medium_side = Side(style='medium', color='64748B')
    header_border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=medium_side)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = header_border

    thin_side = Side(style='thin', color='CBD5E1')
    row_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    for idx, cliente in enumerate(qs, start=2):
        row = [
            cliente.idCliente,
            cliente.nombre or '',
            cliente.rut or '',
            cliente.telefono or '',
            cliente.email or '',
            cliente.direccion or '',
            cliente.estadoCondicion or '',
            cliente.fecha_registro.strftime('%Y-%m-%d %H:%M') if cliente.fecha_registro else '',
        ]
        ws.append(row)

        # Color alternado
        if idx % 2 == 0:
            alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            for c in ws[idx]:
                c.fill = alt_fill

        # Bordes y alineación
        for col_idx, c in enumerate(ws[idx], start=1):
            c.border = row_border
            if col_idx in (1, 7):  # ID y Estado centrados
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(vertical='center')

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 24

    widths = [8, 25, 15, 15, 25, 30, 15, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from datetime import datetime
    filename = f"clientes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ======================================================
# LISTAR CLIENTES (AJAX + PAGINACIÓN)
# ======================================================
def clientes_list(request):
    """Lista de clientes con búsqueda, filtros y paginación (compatible con AJAX parcial)."""
    qs = Cliente.objects.all()

    # --- Búsqueda ---
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q)
            | Q(rut__icontains=q)
            | Q(email__icontains=q)
            | Q(telefono__icontains=q)
            | Q(direccion__icontains=q)
        )

    # --- Orden ---
    sort = (request.GET.get('sort') or 'nombre').strip()
    direction = (request.GET.get('dir') or 'asc').strip().lower()
    sort_map = {
        'nombre': 'nombre',
        'rut': 'rut',
        'email': 'email',
        'telefono': 'telefono',
        'estadoCondicion': 'estadoCondicion',
        'fecha_registro': 'fecha_registro',
    }
    order_field = sort_map.get(sort, 'nombre')
    if direction == 'desc':
        order_field = f'-{order_field}'
    qs = qs.order_by(order_field)

    # --- Tamaño de página ---
    allowed_sizes = [5, 10, 20, 50]
    try:
        page_size = int(request.GET.get('page_size') or request.session.get('cliente_page_size') or 10)
    except ValueError:
        page_size = 10
    if page_size not in allowed_sizes:
        page_size = 10
    request.session['cliente_page_size'] = page_size

    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'clientes': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'q': q,
        'sort': sort,
        'dir': direction,
        'page_size': page_size,
        'page_sizes': allowed_sizes,
    }

    # --- Renderizado parcial (AJAX o Fetch) ---
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    is_partial = request.GET.get('partial') == '1'
    if is_ajax or is_partial:
        return render(request, 'clientes/partials/cliente_table.html', context)

    # --- Renderizado completo ---
    return render(request, 'clientes/cliente_list.html', context)


# ======================================================
# CRUD CLIENTES
# ======================================================
def clientes_create(request):
    """Crear nuevo cliente."""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente creado correctamente.")
            return redirect('clientes_list')
    else:
        form = ClienteForm()
    return render(request, 'clientes/cliente_add.html', {'form': form})


def clientes_detail(request, pk):
    """Ver detalle de un cliente."""
    cliente = get_object_or_404(Cliente, idCliente=pk)
    return render(request, 'clientes/cliente_detail.html', {'cliente': cliente})


def clientes_edit(request, pk):
    """Editar cliente."""
    cliente = get_object_or_404(Cliente, idCliente=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente actualizado correctamente.")
            return redirect('clientes_list')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'clientes/cliente_edit.html', {'form': form, 'cliente': cliente})


def clientes_delete(request, pk):
    """Eliminar cliente."""
    cliente = get_object_or_404(Cliente, idCliente=pk)
    if request.method == 'POST':
        nombre = str(cliente.nombre)
        cliente.delete()
        messages.success(request, f'Cliente "{nombre}" eliminado correctamente.')
        return redirect('clientes_list')
    return redirect('clientes_detail', pk=pk)


# ======================================================
# CLASES LEGACY (opcional, no usadas en AJAX)
# ======================================================
class ClienteListView(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = "clientes/cliente_list.html"
    context_object_name = "clientes"


    def get_queryset(self):
        qs = Cliente.objects.all()
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(rut__icontains=q)
                | Q(email__icontains=q)
                | Q(telefono__icontains=q)
                | Q(direccion__icontains=q)
            )

        sort = (self.request.GET.get("sort") or "nombre").strip()
        direction = (self.request.GET.get("dir") or "asc").strip().lower()
        sort_map = {
            "nombre": "nombre",
            "rut": "rut",
            "email": "email",
            "telefono": "telefono",
            "estadoCondicion": "estadoCondicion",
            "fecha_registro": "fecha_registro",
        }
        order_field = sort_map.get(sort, "nombre")
        if direction == "desc":
            order_field = f"-{order_field}"
        return qs.order_by(order_field)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "")
        context["sort"] = self.request.GET.get("sort", "nombre")
        context["dir"] = self.request.GET.get("dir", "asc")
        context["page_size"] = int(self.request.GET.get("page_size", 10))
        context["page_sizes"] = [5, 10, 20, 50]
        return context


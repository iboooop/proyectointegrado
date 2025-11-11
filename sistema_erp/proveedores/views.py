from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import Proveedor
from productos.models import Producto
from transacciones.models import MovimientoInventario
from .forms import ProveedorForm
from django.core.paginator import Paginator

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    Workbook = None

def lista_proveedores(request):
    qs = Proveedor.objects.all().order_by('nombre')

    # --- Búsqueda ---
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) |
            Q(rut__icontains=q) |
            Q(contacto__icontains=q) |
            Q(telefono__icontains=q) |
            Q(correo__icontains=q) |
            Q(direccion__icontains=q)
        )

    # --- Orden (si lo deseas implementar) ---
    sort = (request.GET.get('sort') or 'nombre').strip()
    direction = (request.GET.get('dir') or 'asc').strip().lower()
    if direction == 'desc':
        sort = f'-{sort}'
    qs = qs.order_by(sort)

    # --- Paginación ---
    allowed_sizes = [5, 10, 20, 50]
    try:
        page_size = int(request.GET.get('page_size') or request.session.get('proveedor_page_size') or 10)
    except ValueError:
        page_size = 10
    if page_size not in allowed_sizes:
        page_size = 10
    request.session['proveedor_page_size'] = page_size

    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'proveedores': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'q': q,
        'sort': sort,
        'dir': direction,
        'page_size': page_size,
        'page_sizes': allowed_sizes,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('partial') == '1':
        return render(request, 'proveedores/partials/proveedor_table.html', context)

    return render(request, 'proveedores/proveedor_list.html', context)


def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm()
    return render(request, 'proveedores/proveedor_add.html', {'form': form})

def detalle_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    productos = Producto.objects.filter(proveedor=proveedor)
    movimientos = MovimientoInventario.objects.filter(proveedor=proveedor).select_related('producto').order_by('-fecha')

    context = {
        'proveedor': proveedor,
        'productos': productos,
        'movimientos': movimientos,
    }
    return render(request, 'proveedores/proveedor_detail.html', context)

def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('detalle_proveedor', id=proveedor.id)
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'proveedores/proveedor_edit.html', {'form': form, 'proveedor': proveedor})

def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    if request.method == 'POST':
        proveedor.delete()
        return redirect('lista_proveedores')
    return redirect('detalle_proveedor', id=id)


def exportar_proveedores_excel(request):
    if Workbook is None:
        return HttpResponse("openpyxl no está instalado.", status=500)

    qs = Proveedor.objects.all().order_by('id')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Proveedores'
    headers = ['ID', 'Nombre', 'RUT', 'Teléfono', 'Estado']
    header_fill = PatternFill(start_color='EAF2FF', end_color='EAF2FF', fill_type='solid')
    bold = Font(bold=True, color='1f2937')
    center = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    medium_side = Side(style='medium', color='64748B')
    header_border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=medium_side)
    for c in ws[1]:
        c.font = bold; c.fill = header_fill; c.alignment = center; c.border = header_border

    thin_side = Side(style='thin', color='CBD5E1')
    row_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    for idx, p in enumerate(qs, start=2):
        row = [p.id, p.nombre, getattr(p, 'rut', ''), getattr(p, 'telefono', ''), getattr(p, 'estado', '')]
        ws.append(row)
        if idx % 2 == 0:
            alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            for c in ws[idx]: c.fill = alt_fill
        for col_i, c in enumerate(ws[idx], start=1):
            c.border = row_border
            if col_i in (1,5): c.alignment = Alignment(horizontal='center', vertical='center')
            else: c.alignment = Alignment(vertical='center')

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    from openpyxl.utils import get_column_letter
    widths = [8, 30, 20, 18, 14]
    for i, w in enumerate(widths, start=1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    for r in range(2, ws.max_row+1): ws.row_dimensions[r].height = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from datetime import datetime
    response['Content-Disposition'] = f'attachment; filename="proveedores_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"'
    wb.save(response)
    return response

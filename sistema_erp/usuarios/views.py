from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from .forms import UsuarioForm, PerfilForm
from .models import Perfil
from productos.models import Producto
from proveedores.models import Proveedor
from transacciones.models import MovimientoInventario
from django.contrib.auth.models import User

from django.http import HttpResponse, HttpResponseForbidden
from django.core.mail import send_mail
from django.conf import settings

from datetime import datetime
import random
import string
import threading

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except Exception:
    Workbook = None


# ---------------- FUNCIONES AUXILIARES ----------------
def generar_password_robusta():
    """
    Genera una contraseña robusta de 12 caracteres que incluye:
    - Al menos 1 mayúscula
    - Al menos 1 minúscula
    - Al menos 1 dígito
    - Al menos 1 carácter especial
    """
    mayusculas = string.ascii_uppercase
    minusculas = string.ascii_lowercase
    digitos = string.digits
    especiales = '!@#$%&*'
    
    # Asegurar al menos uno de cada tipo
    password = [
        random.choice(mayusculas),
        random.choice(minusculas),
        random.choice(digitos),
        random.choice(especiales)
    ]
    
    # Completar hasta 12 caracteres con caracteres aleatorios
    todos = mayusculas + minusculas + digitos + especiales
    password += [random.choice(todos) for _ in range(8)]
    
    # Mezclar para que no sea predecible
    random.shuffle(password)
    return ''.join(password)


def enviar_correo_clave_provisoria(usuario, password_provisoria, es_reset=False):
    """
    Envía un correo al usuario con su contraseña provisoria usando template HTML.
    Usa timeout corto para no bloquear la respuesta.
    
    Args:
        usuario: Objeto User de Django
        password_provisoria: Contraseña temporal generada
        es_reset: True si es un reset de contraseña, False si es creación de usuario
    """
    try:
        from django.template.loader import render_to_string
        from django.utils.html import strip_tags
        from email.mime.image import MIMEImage
        import os
        
        # Determinar el template y asunto según el tipo de correo
        if es_reset:
            template_name = 'autenticacion/emails/reset_password.html'
            asunto = 'Contraseña Restablecida - Lilis ERP'
        else:
            template_name = 'autenticacion/emails/credenciales.html'
            asunto = 'Bienvenido - Credenciales de Acceso - Lilis ERP'
        
        # Contexto para el template
        context = {
            'nombre': usuario.first_name or usuario.username,
            'username': usuario.username,
            'password': password_provisoria,
            'url_login': 'http://127.0.0.1:8000/autenticacion/login/',
        }
        
        # Renderizar el template HTML
        html_message = render_to_string(template_name, context)
        # Crear versión de texto plano
        plain_message = strip_tags(html_message)
        
        # Enviar con timeout de 3 segundos y fail_silently para no bloquear
        from django.core.mail import EmailMultiAlternatives
        email = EmailMultiAlternatives(
            asunto,
            plain_message,
            settings.DEFAULT_FROM_EMAIL,
            [usuario.email]
        )
        email.attach_alternative(html_message, "text/html")
        
        # Adjuntar logo como imagen embebida
        logo_path = os.path.join(settings.BASE_DIR, 'autenticacion', 'static', 'assets', 'logolilis_ver2.png')
        if os.path.exists(logo_path):
            with open(logo_path, 'rb') as f:
                logo_data = f.read()
                logo = MIMEImage(logo_data)
                logo.add_header('Content-ID', '<logo>')
                logo.add_header('Content-Disposition', 'inline', filename='logo.png')
                email.attach(logo)
        
        email.send(fail_silently=True)
        return True
    except Exception as e:
        print(f"Error al enviar correo: {e}")
        return False

# ---------------- DASHBOARD ----------------
@login_required
def dashboard_view(request):
    # obtener perfil y rol de forma segura
    perfil = Perfil.objects.filter(usuario=request.user).first()
    rol = getattr(perfil, 'rol', None)


    # sincronizar rol en la sesión para que la plantilla pueda leerlo
    if rol:
        request.session['rol'] = rol

    else:
        request.session.pop('rol', None)

    # decidir permisos (ADMIN / EDITOR / LECTOR)
    can_view_all = rol in ('ADMIN', 'LECTOR', 'EDITOR')

    # aquí debes rellenar el contexto que ya tenías (totales, listas, etc).
    context = {
        # ... existing context values (ej: total_productos, ultimos_productos, etc) ...
        'can_view_all': can_view_all,
    }
    return render(request, 'sistema_erp/templates/dashboard.html', context)


# ---------------- LISTADO ----------------
@login_required
def usuarios_list_view(request):
    perfil = Perfil.objects.filter(usuario=request.user).first()
    rol = getattr(perfil, 'rol', None)

    # permitir ADMIN, LECTOR y EDITOR; denegar el resto
    if rol not in ('ADMIN', 'LECTOR', 'EDITOR'):
        return HttpResponseForbidden("No tiene permisos para ver usuarios.")
    
    # Limpiar parámetros de query antiguos si no es una búsqueda activa
    if not request.GET.get('q') and request.GET.get('created'):
        # Limpiar el parámetro 'created' redirigiendo sin él
        return redirect('usuarios_list')

    # sincronizar sesión por si se accede directamente
    if rol:
        request.session['rol'] = rol

    # Limpiar parámetros de query antiguos si se accede directamente sin búsqueda
    if not request.GET.get('q') and (request.GET.get('created') or request.GET.get('deleted')):
        # Redirigir sin parámetros antiguos
        return redirect('usuarios_list')

    # Filtros de búsqueda
    q = (request.GET.get('q') or '').strip()

    # Ordenación segura por campos permitidos
    sort = (request.GET.get('sort') or 'usuario__username').strip()
    direction = (request.GET.get('dir') or 'asc').strip().lower()

    sort_map = {
        'usuario__username': 'usuario__username',
        'usuario__email': 'usuario__email',
        'usuario__first_name': 'usuario__first_name',
        'usuario__last_name': 'usuario__last_name',
        'telefono': 'telefono',
        'rol': 'rol',
        'estado': 'estado',
        'mfa_habilitado': 'mfa_habilitado',
        'usuario__last_login': 'usuario__last_login',
        'sesiones_activas': 'sesiones_activas',
    }

    base_qs = Perfil.objects.select_related('usuario').all()
    if q:
        base_qs = base_qs.filter(
            Q(usuario__first_name__icontains=q)
            | Q(usuario__last_name__icontains=q)
            | Q(usuario__username__icontains=q)
            | Q(usuario__email__icontains=q)
            | Q(telefono__icontains=q)
        )

    order_field = sort_map.get(sort, 'usuario__username')
    if direction == 'desc':
        order_field = f'-{order_field}'
    base_qs = base_qs.order_by(order_field)

    # Paginación
    # Tamaño de página configurable (persistente en sesión)
    allowed_page_sizes = [5, 10, 20, 50, 100]
    try:
        page_size = int(request.GET.get('page_size') or request.session.get('usuarios_page_size') or 10)
    except ValueError:
        page_size = 10
    if page_size not in allowed_page_sizes:
        page_size = 10
    request.session['usuarios_page_size'] = page_size

    paginator = Paginator(base_qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'perfiles': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'q': q,
        'sort': sort,
        'dir': direction,
        'page_size': page_size,
        'page_sizes': allowed_page_sizes,
    }

    # Respuesta parcial para AJAX (solo tabla y paginación)
    if request.GET.get('partial') == '1' or request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'usuarios/partials/list_table.html', context)

    return render(request, 'usuarios/list.html', context)


# ---------------- CREAR ----------------
@login_required
def usuarios_create_view(request):
    perfil_admin = Perfil.objects.filter(usuario=request.user).first()
    rol_admin = getattr(perfil_admin, 'rol', None)
    
    # Solo ADMIN puede crear usuarios
    if rol_admin != 'ADMIN':
        return HttpResponseForbidden("Solo los administradores pueden crear usuarios.")
    
    if request.method == 'POST':
        usuario_form = UsuarioForm(request.POST)
        perfil_form = PerfilForm(request.POST, request.FILES)

        if usuario_form.is_valid() and perfil_form.is_valid():
            try:
                with transaction.atomic():
                    # Generar contraseña provisoria robusta
                    password_provisoria = generar_password_robusta()
                    
                    # Crear usuario sin contraseña primero
                    usuario = usuario_form.save(commit=False)
                    usuario.set_password(password_provisoria)
                    usuario.save()
                    
                    # Crear perfil asociado
                    perfil = perfil_form.save(commit=False)
                    perfil.usuario = usuario
                    perfil.debe_cambiar_clave = True  # Marcar para cambio obligatorio
                    perfil.save()
                    
                # Enviar correo en segundo plano para no bloquear la respuesta
                def enviar_correo_background():
                    try:
                        enviar_correo_clave_provisoria(usuario, password_provisoria)
                    except:
                        pass  # Silenciar errores en el thread
                
                thread = threading.Thread(target=enviar_correo_background)
                thread.daemon = True  # El thread se cerrará cuando termine el proceso principal
                thread.start()
                
                # Mostrar mensaje de éxito inmediatamente (sin esperar el correo)
                messages.success(request, f"Usuario '{usuario.username}' creado exitosamente. Se enviará un correo a {usuario.email} con las credenciales de acceso.")
                
                # Redirigir al listado (el mensaje se mostrará allí)
                return redirect('usuarios_list')
            except Exception as e:
                messages.error(request, f"Hubo un error al crear el usuario: {str(e)}")
        else:
            print("Errores UsuarioForm:", usuario_form.errors)
            print("Errores PerfilForm:", perfil_form.errors)
            messages.error(request, "Hubo un error al crear el usuario. Verifique los datos ingresados.")
    else:
        # 🔹 Limpiar TODOS los mensajes antiguos al cargar en GET
        # Esto evita que mensajes de otras vistas aparezcan aquí
        list(messages.get_messages(request))  # Consumir y descartar todos los mensajes
        usuario_form = UsuarioForm()
        perfil_form = PerfilForm()

    return render(request, 'usuarios/create.html', {
        'usuario_form': usuario_form,
        'perfil_form': perfil_form,
    })


# ---------------- EDITAR ----------------
@login_required
def usuarios_edit_view(request, id):
    perfil = get_object_or_404(Perfil, id=id)
    usuario = perfil.usuario
    updated = False  # 🔹 Bandera para saber si se guardó correctamente

    if request.method == 'POST':
        usuario_form = UsuarioForm(request.POST, instance=usuario)
        perfil_form = PerfilForm(request.POST, request.FILES, instance=perfil)

        if usuario_form.is_valid() and perfil_form.is_valid():
            # Detectar si realmente hubo cambios comparando datos limpios con la BD
            user_changed = False
            perfil_changed = False
            
            # Debug: imprimir changed_data
            print("DEBUG - usuario_form.changed_data:", usuario_form.changed_data)
            print("DEBUG - perfil_form.changed_data:", perfil_form.changed_data)
            
            # Verificar cambios en usuario (ignorar campos que siempre cambian)
            for field in usuario_form.changed_data:
                if field not in ['password']:  # Ignorar campos que no importan
                    user_changed = True
                    break
            
            # Verificar cambios en perfil (ignorar sesiones_activas si es el único cambio)
            perfil_real_changes = [f for f in perfil_form.changed_data if f not in ['sesiones_activas']]
            if perfil_real_changes:
                perfil_changed = True

            print("DEBUG - user_changed:", user_changed)
            print("DEBUG - perfil_changed:", perfil_changed)

            if not (user_changed or perfil_changed):
                messages.info(request, "No se detectaron cambios para guardar.")
                # Redirigir para limpiar el POST
                return redirect(reverse('usuarios_edit', args=[perfil.id]))

            # Guardar solo si hubo cambios
            usuario_form.save()
            perfil_form.save()
            # Mensaje flash de éxito (consumido una sola vez)
            messages.success(request, "Los cambios se han guardado correctamente.")
            # Redirigir sin query params para evitar re-mostrar al refrescar
            return redirect(reverse('usuarios_edit', args=[perfil.id]))

        else:
            print("Errores UsuarioForm:", usuario_form.errors)
            print("Errores PerfilForm:", perfil_form.errors)
            messages.error(request, "No se pudieron guardar los cambios. Verifique los datos.")
    else:
        # 🔹 Limpiar TODOS los mensajes antiguos al cargar en GET
        list(messages.get_messages(request))  # Consumir y descartar todos los mensajes
        usuario_form = UsuarioForm(instance=usuario)
        perfil_form = PerfilForm(instance=perfil)

    return render(request, 'usuarios/edit.html', {
        'perfil': perfil,
        'usuario_form': usuario_form,
        'perfil_form': perfil_form,
    })


# ---------------- DETALLE ----------------
@login_required
def usuarios_detail_view(request, id):
    perfil = get_object_or_404(Perfil, id=id)
    
    # Verificar si se reseteo la contraseña exitosamente
    show_reset_alert = request.session.pop('password_reset_success', False)
    reset_email = request.session.pop('password_reset_email', '')
    
    return render(request, 'usuarios/details.html', {
        'perfil': perfil,
        'show_reset_alert': show_reset_alert,
        'reset_email': reset_email,
    })


# ---------------- ELIMINAR ----------------
@login_required
def usuarios_delete_view(request, id):
    perfil = get_object_or_404(Perfil, id=id)
    usuario = perfil.usuario

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Eliminar primero el perfil y luego el usuario
                perfil.delete()
                usuario.delete()
            messages.success(request, "Usuario eliminado correctamente.")
        except Exception as e:
            messages.error(request, f"Error al eliminar el usuario: {str(e)}")
        return redirect('usuarios_list')
    # ... lógica para mostrar confirmación si es GET ...

    return render(request, 'usuarios/delete.html', {'perfil': perfil})


# ---------------- EXPORTAR EXCEL ----------------
@login_required
def export_usuarios_excel(request):
    """Exporta todos los usuarios (User + Perfil) a un archivo XLSX."""
    if Workbook is None:
        return HttpResponse("openpyxl no está instalado. Agrega 'openpyxl' a requirements.txt e instala las dependencias.", status=500)

    perfiles = Perfil.objects.select_related('usuario').order_by('usuario__username')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Usuarios'

    headers = [
        'Username', 'Email', 'Nombres', 'Apellidos', 'Teléfono',
        'Rol', 'Estado', 'MFA', 'Último acceso', 'Sesiones activas'
    ]
    header_fill = PatternFill(start_color='EAF2FF', end_color='EAF2FF', fill_type='solid')
    bold = Font(bold=True, color='1f2937')
    center = Alignment(horizontal='center', vertical='center')

    ws.append(headers)
    # Estilos de encabezado con bordes gruesos
    from openpyxl.styles import Border, Side
    medium_side = Side(style='medium', color='64748B')
    header_border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=medium_side)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = header_border

    thin_side = Side(style='thin', color='CBD5E1')
    row_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    for idx, p in enumerate(perfiles, start=2):
        u = p.usuario
        row = [
            u.username,
            u.email,
            u.first_name,
            u.last_name,
            p.telefono or '',
            p.get_rol_display() if hasattr(p, 'get_rol_display') else p.rol,
            p.estado,
            'Sí' if p.mfa_habilitado else 'No',
            u.last_login.strftime('%Y-%m-%d %H:%M') if u.last_login else '',
            p.sesiones_activas,
        ]
        ws.append(row)
        # Alternancia de filas para mejor lectura
        # Alternancia filas pares
        if idx % 2 == 0:
            alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            for c in ws[idx]:
                c.fill = alt_fill
        # Bordes y alineaciones
        for col_i, c in enumerate(ws[idx], start=1):
            c.border = row_border
            if col_i in (1,2,3,4,6,7):  # texto clave
                c.alignment = Alignment(vertical='center')
            if col_i == 9:  # Último acceso
                c.alignment = Alignment(horizontal='center', vertical='center')
            if col_i == 10:  # Sesiones activas
                c.alignment = Alignment(horizontal='center', vertical='center')

    # Formatos y UX de hoja: auto-filter, freeze panes y anchos
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    # Ajustar altura encabezado y datos
    ws.row_dimensions[1].height = 24
    for r in range(2, ws.max_row+1):
        ws.row_dimensions[r].height = 18
    from openpyxl.utils import get_column_letter
    widths = [18, 28, 20, 20, 16, 18, 14, 10, 20, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"usuarios_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ---------------- RESETEAR CONTRASEÑA (RQ-USR-06) ----------------
@login_required
def usuarios_reset_password_view(request, id):
    """RQ-USR-06: El administrador puede resetear la contraseña de un usuario.
    Se genera una nueva clave temporal robusta y se envía por correo.
    """
    perfil_admin = Perfil.objects.filter(usuario=request.user).first()
    rol_admin = getattr(perfil_admin, 'rol', None)
    
    # Solo ADMIN puede resetear contraseñas
    if rol_admin != 'ADMIN':
        return HttpResponseForbidden("Solo los administradores pueden resetear contraseñas.")
    
    perfil = get_object_or_404(Perfil, id=id)
    usuario = perfil.usuario
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Generar nueva contraseña provisoria robusta
                password_provisoria = generar_password_robusta()
                
                # Actualizar contraseña del usuario
                usuario.set_password(password_provisoria)
                usuario.save()
                
                # Marcar que debe cambiar la clave
                perfil.debe_cambiar_clave = True
                perfil.save()
                
            # Enviar correo en segundo plano con el template de reset
            def enviar_correo_background():
                try:
                    enviar_correo_clave_provisoria(usuario, password_provisoria, es_reset=True)
                except:
                    pass  # Silenciar errores en el thread
            
            thread = threading.Thread(target=enviar_correo_background)
            thread.daemon = True
            thread.start()
            
            # Mensaje de éxito con el email del usuario
            messages.success(
                request, 
                f"Contraseña restablecida. Se ha enviado un correo a {usuario.email} con la nueva contraseña provisional. El usuario deberá cambiarla obligatoriamente en su próximo inicio de sesión."
            )
            return redirect(reverse('usuarios_detail', args=[perfil.id]))
        except Exception as e:
            messages.error(request, f"Error al resetear la contraseña: {str(e)}")
            return redirect(reverse('usuarios_detail', args=[perfil.id]))
    
    # Si es GET, mostrar confirmación
    return render(request, 'usuarios/reset_password_confirm.html', {'perfil': perfil})
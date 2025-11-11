from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.contrib import messages
from django.views.generic import CreateView, UpdateView, DetailView, View
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from django import forms

from .models import Bodega

# ============================================
# FORMULARIO
# ============================================
try:
    from .forms import BodegaForm  # type: ignore
except Exception:
    class BodegaForm(forms.ModelForm):
        class Meta:
            model = Bodega
            fields = [
                "codigo",
                "nombre",
                "direccion",
                "telefono",
                "responsable",
                "capacidad_maxima",
                "tipo",
                "estado",
            ]


# ============================================
# LISTADO CON BÚSQUEDA, FILTROS Y PAGINACIÓN
# ============================================
def bodegas_list(request):
    """Lista de bodegas con búsqueda, orden y paginación (compatible con AJAX parcial)"""
    qs = Bodega.objects.all()

    # --- Búsqueda ---
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q)
            | Q(codigo__icontains=q)
            | Q(direccion__icontains=q)
            | Q(responsable__icontains=q)
            | Q(tipo__icontains=q)
            | Q(estado__icontains=q)
        )

    # --- Orden ---
    sort = (request.GET.get('sort') or 'nombre').strip()
    direction = (request.GET.get('dir') or 'asc').strip().lower()
    sort_map = {
        'codigo': 'codigo',
        'nombre': 'nombre',
        'direccion': 'direccion',
        'responsable': 'responsable',
        'tipo': 'tipo',
        'estado': 'estado',
    }
    order_field = sort_map.get(sort, 'nombre')
    if direction == 'desc':
        order_field = f'-{order_field}'
    qs = qs.order_by(order_field)

    # --- Tamaño de página persistente ---
    allowed_sizes = [5, 10, 20, 50]
    try:
        page_size = int(request.GET.get('page_size') or request.session.get('bodega_page_size') or 10)
    except ValueError:
        page_size = 10
    if page_size not in allowed_sizes:
        page_size = 10
    request.session['bodega_page_size'] = page_size

    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'bodegas': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'q': q,
        'sort': sort,
        'dir': direction,
        'page_size': page_size,
        'page_sizes': allowed_sizes,
    }

    # --- Render parcial si es AJAX ---
    if request.GET.get('partial') == '1' or request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'bodegas/partials/bodega_table.html', context)

    # --- Render completo ---
    return render(request, 'bodegas/bodega_list.html', context)


# ============================================
# CREAR / EDITAR / DETALLE / ELIMINAR
# ============================================
class BodegaCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Bodega
    form_class = BodegaForm
    template_name = "bodegas/bodega_add.html"
    success_url = reverse_lazy("bodegas_list")
    success_message = "Bodega creada correctamente."


class BodegaUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Bodega
    form_class = BodegaForm
    template_name = "bodegas/bodega_edit.html"
    success_url = reverse_lazy("bodegas_list")
    success_message = "Bodega actualizada correctamente."


class BodegaDetailView(LoginRequiredMixin, DetailView):
    model = Bodega
    template_name = "bodegas/bodega_detail.html"
    context_object_name = "bodega"


class BodegaDeleteView(LoginRequiredMixin, View):
    """Eliminación vía POST"""
    def post(self, request, pk=None, *args, **kwargs):
        bodega = get_object_or_404(Bodega, pk=pk)
        nombre = str(bodega.nombre)
        bodega.delete()
        messages.success(request, f'Bodega "{nombre}" eliminada correctamente.')
        return redirect(reverse_lazy("bodegas_list"))


# ============================================
# EXPORTAR A EXCEL
# ============================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


def exportar_bodegas_excel(request):
    """Exporta bodegas filtradas/ordenadas a un archivo XLSX."""
    if Workbook is None:
        return HttpResponse("openpyxl no está instalado.", status=500)

    qs = Bodega.objects.all()

    # Filtros
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q)
            | Q(codigo__icontains=q)
            | Q(direccion__icontains=q)
            | Q(responsable__icontains=q)
            | Q(tipo__icontains=q)
            | Q(estado__icontains=q)
        )

    # Orden
    sort = (request.GET.get('sort') or 'nombre').strip()
    direction = (request.GET.get('dir') or 'asc').strip().lower()
    sort_map = {
        'codigo': 'codigo',
        'nombre': 'nombre',
        'direccion': 'direccion',
        'responsable': 'responsable',
        'tipo': 'tipo',
        'estado': 'estado',
    }
    order_field = sort_map.get(sort, 'nombre')
    if direction == 'desc':
        order_field = f'-{order_field}'
    qs = qs.order_by(order_field)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bodegas'

    headers = ['Código', 'Nombre', 'Dirección', 'Responsable', 'Tipo', 'Estado']
    header_fill = PatternFill(start_color='EAF2FF', end_color='EAF2FF', fill_type='solid')
    bold = Font(bold=True, color='1f2937')
    center = Alignment(horizontal='center', vertical='center')

    ws.append(headers)
    medium_side = Side(style='medium', color='64748B')
    header_border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=medium_side)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = header_border

    thin_side = Side(style='thin', color='CBD5E1')
    row_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    for idx, b in enumerate(qs, start=2):
        row = [
            b.codigo or '',
            b.nombre or '',
            b.direccion or '',
            b.responsable or '',
            b.tipo or '',
            b.estado or '',
        ]
        ws.append(row)

        if idx % 2 == 0:
            alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            for c in ws[idx]:
                c.fill = alt_fill
        for c in ws[idx]:
            c.border = row_border
            c.alignment = Alignment(vertical='center')

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 24

    widths = [15, 25, 35, 25, 20, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from datetime import datetime
    filename = f"bodegas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response

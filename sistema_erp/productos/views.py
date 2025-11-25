from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .models import Producto
from .forms import ProductoForm
from transacciones.models import MovimientoInventario
from django.core.paginator import Paginator
from django.db.models import Q, Avg

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


# ---------------- LISTAR ----------------
@login_required
def lista_productos(request):

    qs = Producto.objects.select_related("proveedor").all()

    q = (request.GET.get("q") or "").strip()
    if q:
        # Buscar solo en campos de texto; tratar "activo"/"inactivo" por separado
        q_lower = q.lower()
        filters = Q(nombre__icontains=q) | Q(sku__icontains=q) | Q(proveedor__nombre__icontains=q) | Q(categoria__icontains=q)
        if q_lower in ("activo", "activos"):
            filters |= Q(activo=True)
        elif q_lower in ("inactivo", "inactivos"):
            filters |= Q(activo=False)
        qs = qs.filter(filters)

    sort = (request.GET.get("sort") or "nombre").strip()
    direction = (request.GET.get("dir") or "asc").strip().lower()

    allowed_sorts = {
        "sku": "sku",
        "nombre": "nombre",
        "stock": "stock_actual",
        "estado": "activo",
    }
    sort_field = allowed_sorts.get(sort, "nombre")
    if direction == "desc":
        sort_field = f"-{sort_field}"
    qs = qs.order_by(sort_field)

    # tamaños permitidos coherentes con los selects del template
    allowed_sizes = [10, 25, 50, 100]
    try:
        page_size = int(
            request.GET.get("page_size") or request.session.get("producto_page_size") or 10
        )
    except ValueError:
        page_size = 10
    if page_size not in allowed_sizes:
        page_size = 10
    request.session["producto_page_size"] = page_size

    paginator = Paginator(qs, page_size)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "productos": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "q": q,
        "sort": sort,
        "dir": direction,
        "page_size": page_size,
        "page_sizes": allowed_sizes,
    }

    if (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or request.GET.get("partial") == "1"
    ):
        return render(request, "productos/partials/producto_table.html", context)

    return render(request, "productos/product_list.html", context)



# ---------------- CREAR ----------------
@login_required
def crear_producto(request):
    active_tab = request.POST.get("active_tab", "paso1-tab")
    mensaje = None
    mensaje_tipo = None

    if request.method == "POST":
        data = request.POST.copy()

        sku_letras = (data.get("sku_letras") or "").strip()
        sku_nros = (data.get("sku_nros") or "").strip()
        if sku_letras or sku_nros:
            data["sku"] = f"{sku_letras}{sku_nros}"

        form = ProductoForm(data, files=request.FILES)
        if form.is_valid():
            form.save()
            mensaje = "Producto creado exitosamente."
            mensaje_tipo = "success"
            form = ProductoForm()
            active_tab = "paso1-tab"
        else:
            mensaje = "Corrige los errores indicados."
            mensaje_tipo = "danger"
    else:
        form = ProductoForm()

    return render(
        request,
        "productos/product_add.html",
        {
            "form": form,
            "active_tab": active_tab,
            "mensaje": mensaje,
            "mensaje_tipo": mensaje_tipo,
        },
    )


# ---------------- DETALLE ----------------
@login_required
def detalle_producto(request, id):
    producto = get_object_or_404(
        Producto.objects.select_related("proveedor"), idProducto=id
    )
    movimientos = (
        MovimientoInventario.objects.filter(producto=producto)
        .select_related("proveedor", "usuario")
        .order_by("-fecha")
    )
    # promedio del costo_estandar entre productos de la misma categoría
    costo_promedio = (
        Producto.objects.filter(categoria=producto.categoria)
        .aggregate(avg_costo=Avg('costo_estandar'))
        .get('avg_costo')
    )

    return render(
        request,
        "productos/product_detail.html",
        {
            "producto": producto,
            "movimientos": movimientos,
            "costo_promedio_categoria": costo_promedio,
        },
    )


# ---------------- EDITAR ----------------
@login_required
def editar_producto(request, id):
    producto = get_object_or_404(Producto, idProducto=id)

    mensaje = None
    mensaje_tipo = None

    if request.method == "POST":
        data = request.POST.copy()

        sku_letras = (data.get("sku_letras") or "").strip()
        sku_nros = (data.get("sku_nros") or "").strip()
        if sku_letras or sku_nros:
            data["sku"] = f"{sku_letras}{sku_nros}"

        form = ProductoForm(data, files=request.FILES, instance=producto)
        if form.is_valid():
            if not form.has_changed():
                # No hay cambios reales en los datos del formulario
                mensaje = "No realizaste ningún cambio."
                mensaje_tipo = "warning"
            else:
                from django.utils import timezone

                was_inactive = not producto.activo
                producto = form.save(commit=False)
                now = timezone.now()

                if producto.activo and was_inactive:
                    producto.fecha_activacion = now
                    producto.fecha_desactivacion = None
                elif not producto.activo and was_inactive is False:
                    producto.fecha_desactivacion = now

                producto.save()
                mensaje = "Cambios guardados correctamente."
                mensaje_tipo = "success"
        else:
            mensaje = "Corrige los errores indicados."
            mensaje_tipo = "danger"
    else:
        # separar sku en letras/nros si quieres reutilizar los 2 inputs
        initial = {
            "sku_letras": producto.sku[:-4],
            "sku_nros": producto.sku[-4:],
        }
        form = ProductoForm(instance=producto, initial=initial)

    return render(
        request,
        "productos/product_edit.html",
        {"form": form, "producto": producto, "mensaje": mensaje, "mensaje_tipo": mensaje_tipo},
    )


# ---------------- ELIMINAR ----------------
@login_required
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, idProducto=id)
    if request.method == "POST":
        from django.utils import timezone

        producto.activo = False
        producto.fecha_desactivacion = timezone.now()
        # No tocamos fecha_activacion aquí para conservar el histórico de alta
        producto.save(update_fields=["activo", "fecha_desactivacion"])
        return redirect("lista_productos")
    return redirect("detalle_producto", id=id)


# ---------------- EXPORTAR EXCEL ----------------
@login_required
def exportar_productos_excel(request):
    if Workbook is None:
        return HttpResponse("openpyxl no está instalado.", status=500)

    # Aplicar los mismos filtros de búsqueda y orden que en lista_productos
    qs = Producto.objects.select_related("proveedor").all()

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q)
            | Q(sku__icontains=q)
            | Q(proveedor__nombre__icontains=q)
            | Q(stock_actual__icontains=q)
            | Q(activo__icontains=(q.lower() in ["activo", "activos", "inactivo", "inactivos"]))
        )

    sort = (request.GET.get("sort") or "nombre").strip()
    direction = (request.GET.get("dir") or "asc").strip().lower()
    allowed_sorts = {
        "sku": "sku",
        "nombre": "nombre",
        "stock": "stock_actual",
        "estado": "activo",
    }
    sort_field = allowed_sorts.get(sort, "nombre")
    if direction == "desc":
        sort_field = f"-{sort_field}"
    qs = qs.order_by(sort_field)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'

    # Encabezados alineados con las columnas principales de la grilla
    headers = [
        'ID',
        'SKU',
        'Nombre',
        'Categoría',
        'Proveedor',
        'Stock actual',
        'Estado',
        'Fecha activación',
        'Fecha desactivación',
    ]
    header_fill = PatternFill(start_color='EAF2FF', end_color='EAF2FF', fill_type='solid')
    bold = Font(bold=True, color='1f2937')
    center = Alignment(horizontal='center', vertical='center')

    ws.append(headers)
    medium_side = Side(style='medium', color='64748B')
    header_border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=medium_side)
    for c in ws[1]:
        c.font = bold
        c.fill = header_fill
        c.alignment = center
        c.border = header_border

    thin_side = Side(style='thin', color='CBD5E1')
    row_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    for idx, p in enumerate(qs, start=2):
        row = [
            p.idProducto,
            p.sku,
            p.nombre,
            p.categoria,
            getattr(p.proveedor, 'nombre', '') if p.proveedor else '',
            p.stock_actual,
            'Activo' if p.activo else 'Inactivo',
            p.fecha_activacion.strftime('%Y-%m-%d %H:%M') if p.fecha_activacion else '',
            p.fecha_desactivacion.strftime('%Y-%m-%d %H:%M') if p.fecha_desactivacion else '',
        ]
        ws.append(row)

        # Filas alternadas
        if idx % 2 == 0:
            alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            for c in ws[idx]:
                c.fill = alt_fill

        for c in ws[idx]:
            c.border = row_border
            c.alignment = Alignment(vertical='center')

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 24

    widths = [8, 14, 26, 18, 24, 12, 12, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from datetime import datetime
    response['Content-Disposition'] = f'attachment; filename="productos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response

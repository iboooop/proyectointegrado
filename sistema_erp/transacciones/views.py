from django.shortcuts import render, get_object_or_404, redirect
from django.db import models
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

from .models import MovimientoInventario
from .forms import MovimientoInventarioForm



# ---------------- LISTA ----------------
@login_required
def lista_transacciones(request):
    qs = MovimientoInventario.objects.select_related(
        'producto', 'proveedor', 'usuario', 'bodega_origen', 'bodega_destino'
    )

    # Filtros de texto (SKU/nombre producto, proveedor, usuario, tipo, bodega, lote, serie, doc ref, motivo)
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(producto__sku__icontains=q)
            | Q(producto__nombre__icontains=q)
            | Q(proveedor__nombre__icontains=q)
            | Q(usuario__username__icontains=q)
            | Q(tipo__icontains=q)
            | Q(bodega_origen__nombre__icontains=q)
            | Q(bodega_origen__codigo__icontains=q)
            | Q(bodega_destino__nombre__icontains=q)
            | Q(bodega_destino__codigo__icontains=q)
            | Q(lote__icontains=q)
            | Q(serie__icontains=q)
            | Q(fecha_vencimiento__icontains=q)
            | Q(doc_referencia__icontains=q)
            | Q(motivo__icontains=q)
        )

    # Filtros específicos por fecha y tipo (para los pill-filters)
    fecha_filtro = (request.GET.get('fecha') or '').strip()
    if fecha_filtro:
        qs = qs.filter(fecha__date=fecha_filtro)

    tipo_filtro = (request.GET.get('tipo') or '').strip()
    if tipo_filtro:
        qs = qs.filter(tipo=tipo_filtro)

    # Orden
    sort = (request.GET.get('sort') or 'fecha').strip()
    direction = (request.GET.get('dir') or 'desc').strip().lower()
    sort_map = {
        'fecha': 'fecha',
        'tipo': 'tipo',
        'producto__nombre': 'producto__nombre',
        'proveedor__nombre': 'proveedor__nombre',
        'bodega_origen__nombre': 'bodega_origen__nombre',
        'bodega_destino__nombre': 'bodega_destino__nombre',
        'cantidad': 'cantidad',
        'usuario__username': 'usuario__username',
        'lote': 'lote',
        'serie': 'serie',
        'fecha_vencimiento': 'fecha_vencimiento',
        'doc_referencia': 'doc_referencia',
    }
    order_field = sort_map.get(sort, 'fecha')
    if direction == 'desc':
        order_field = f'-{order_field}'
    qs = qs.order_by(order_field)

    # Resumen superior (día actual)
    from django.utils import timezone

    hoy = timezone.localdate()
    # Contar TODOS los movimientos de hoy, no solo los del queryset filtrado
    movimientos_hoy = MovimientoInventario.objects.filter(fecha__date=hoy).count()
    stock_total = qs.aggregate(total=models.Sum('cantidad'))['total'] or 0
    productos_unicos = qs.values('producto__sku').distinct().count()

    # Tamaño de página persistente
    allowed_sizes = [5, 10, 20, 50]
    try:
        page_size = int(request.GET.get('page_size') or request.session.get('trans_page_size') or 10)
    except ValueError:
        page_size = 10
    if page_size not in allowed_sizes:
        page_size = 10
    request.session['trans_page_size'] = page_size

    paginator = Paginator(qs, page_size)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'transacciones': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'q': q,
        'sort': sort,
        'dir': direction,
        'page_size': page_size,

        'page_sizes': allowed_sizes,
        'fecha_filtro': fecha_filtro,
        'tipo_filtro': tipo_filtro,
        'movimientos_hoy': movimientos_hoy,
        'stock_total': stock_total,
        'productos_unicos': productos_unicos,
    }

    if request.GET.get('partial') == '1' or request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'transacciones/partials/transaccion_table.html', context)

    return render(request, 'transacciones/transaccion_list.html', context)


# ---------------- CREAR ----------------
@login_required
def crear_transaccion(request):
    mensaje = None
    mensaje_tipo = None

    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST)
        if form.is_valid():
            mov = form.save(commit=False)
            # Por defecto, todo nuevo movimiento parte "por confirmar"
            if not mov.estado:
                mov.estado = 'POR_CONFIRMAR'
            mov.save()
            form = MovimientoInventarioForm()
            mensaje = "Movimiento creado correctamente."
            mensaje_tipo = "success"
        else:
            mensaje = "Corrige los errores indicados."
            mensaje_tipo = "danger"
    else:
        initial = {'fecha': timezone.now().strftime('%Y-%m-%dT%H:%M')}
        form = MovimientoInventarioForm(initial=initial)

    return render(
        request,
        'transacciones/transaccion_add.html',
        {
            'form': form,
            'mensaje': mensaje,
            'mensaje_tipo': mensaje_tipo,
        },
    )


# ---------------- DETALLE ----------------
@login_required
def detalle_transaccion(request, id):
    transaccion = get_object_or_404(
        MovimientoInventario.objects.select_related(
            'producto', 'proveedor', 'usuario', 'bodega_origen', 'bodega_destino'
        ), 
        id=id
    )
    return render(request, 'transacciones/transaccion_detail.html', {'transaccion': transaccion})


# ---------------- EDITAR ----------------
@login_required
def editar_transaccion(request, id):
    transaccion = get_object_or_404(MovimientoInventario, id=id)

    mensaje = None
    mensaje_tipo = None

    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST, instance=transaccion)
        if form.is_valid():
            if not form.has_changed():
                mensaje = "No realizaste ningún cambio."
                mensaje_tipo = "warning"
            else:
                from django.utils import timezone

                was_inactive = (transaccion.estado in ['CANCELADO', 'DESACTIVADO'])
                mov = form.save(commit=False)
                now = timezone.now()

                if mov.estado == 'EN_PROCESO' and was_inactive:
                    mov.fecha_activacion = now
                    mov.fecha_desactivacion = None
                elif mov.estado in ['CANCELADO', 'DESACTIVADO'] and not was_inactive:
                    mov.fecha_desactivacion = now

                mov.save()
                mensaje = "Cambios guardados correctamente."
                mensaje_tipo = "success"
        else:
            mensaje = "Corrige los errores indicados."
            mensaje_tipo = "danger"
    else:
        initial = {
            'fecha': transaccion.fecha.strftime('%Y-%m-%dT%H:%M')
            if transaccion.fecha else timezone.now().strftime('%Y-%m-%dT%H:%M')
        }
        form = MovimientoInventarioForm(instance=transaccion, initial=initial)

    return render(
        request,
        'transacciones/transaccion_edit.html',
        {
            'form': form,
            'transaccion': transaccion,
            'mensaje': mensaje,
            'mensaje_tipo': mensaje_tipo,
        },
    )


# ---------------- ELIMINAR ----------------
@login_required
def eliminar_transaccion(request, id):
    transaccion = get_object_or_404(MovimientoInventario, id=id)
    if request.method == 'POST':
        from django.utils import timezone

        transaccion.estado = 'DESACTIVADO'
        transaccion.fecha_desactivacion = timezone.now()
        transaccion.save(update_fields=['estado', 'fecha_desactivacion'])
        return redirect('lista_transacciones')
    return redirect('detalle_transaccion', id=id)


# ---------------- EXPORTAR EXCEL ----------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


@login_required
def exportar_transacciones_excel(request):
    """Exporta los movimientos filtrados/ordenados a un archivo XLSX con detalles."""
    if Workbook is None:
        return HttpResponse(
            "openpyxl no está instalado. Agrega 'openpyxl' a requirements.txt e instala las dependencias.",
            status=500
        )

    qs = MovimientoInventario.objects.select_related(
        'producto', 'proveedor', 'usuario', 'bodega_origen', 'bodega_destino'
    )

    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(producto__nombre__icontains=q)
            | Q(proveedor__nombre__icontains=q)
            | Q(usuario__username__icontains=q)
            | Q(tipo__icontains=q)
            | Q(bodega_origen__nombre__icontains=q)
            | Q(bodega_origen__codigo__icontains=q)
            | Q(bodega_destino__nombre__icontains=q)
            | Q(bodega_destino__codigo__icontains=q)
            | Q(lote__icontains=q)
            | Q(serie__icontains=q)
            | Q(doc_referencia__icontains=q)
            | Q(motivo__icontains=q)
        )

    sort = (request.GET.get('sort') or 'fecha').strip()
    direction = (request.GET.get('dir') or 'desc').strip().lower()
    sort_map = {
        'fecha': 'fecha',
        'tipo': 'tipo',
        'producto__nombre': 'producto__nombre',
        'proveedor__nombre': 'proveedor__nombre',
        'bodega_origen__nombre': 'bodega_origen__nombre',
        'bodega_destino__nombre': 'bodega_destino__nombre',
        'cantidad': 'cantidad',
        'usuario__username': 'usuario__username',
        'lote': 'lote',
        'serie': 'serie',
        'fecha_vencimiento': 'fecha_vencimiento',
        'doc_referencia': 'doc_referencia',
    }
    order_field = sort_map.get(sort, 'fecha')
    if direction == 'desc':
        order_field = f'-{order_field}'
    qs = qs.order_by(order_field)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimientos'

    headers = [
        'Fecha', 'Tipo', 'Producto', 'Proveedor', 'Bodega Origen', 'Bodega Destino', 
        'Usuario', 'Cantidad', 'Lote', 'Serie', 'Vence', 'Doc Ref', 'Motivo', 'Observaciones'
    ]
    header_fill = PatternFill(start_color='EAF2FF', end_color='EAF2FF', fill_type='solid')
    bold = Font(bold=True, color='1f2937')
    center = Alignment(horizontal='center', vertical='center')

    ws.append(headers)
    medium_side = Side(style='medium', color='64748B')
    header_border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=medium_side)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = header_border

    thin_side = Side(style='thin', color='CBD5E1')
    row_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    for idx, m in enumerate(qs, start=2):
        row = [
            m.fecha.strftime('%Y-%m-%d %H:%M') if m.fecha else '',
            dict(MovimientoInventario.TIPO_MOVIMIENTO).get(m.tipo, m.tipo),
            getattr(m.producto, 'nombre', ''),
            getattr(m.proveedor, 'nombre', '') if m.proveedor else '',
            getattr(m.bodega_origen, 'nombre', '') if m.bodega_origen else '',
            getattr(m.bodega_destino, 'nombre', '') if m.bodega_destino else '',
            getattr(m.usuario, 'username', '') if m.usuario else '',
            m.cantidad,
            m.lote or '',
            m.serie or '',
            m.fecha_vencimiento.strftime('%Y-%m-%d') if m.fecha_vencimiento else '',
            m.doc_referencia or '',
            m.motivo or '',
            (m.observaciones or '')[:1000],
        ]
        ws.append(row)

        # Color alternado
        if idx % 2 == 0:
            alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            for c in ws[idx]:
                c.fill = alt_fill

        # Bordes y alineación
        for c in ws[idx]:
            c.border = row_border
            c.alignment = Alignment(vertical='center')

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 24

    widths = [18, 12, 28, 22, 18, 10, 16, 16, 14, 18, 14, 16, 24, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from datetime import datetime
    filename = f"movimientos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response
